from openpyxl import load_workbook
import os

file_path = 'c:\\Desarrollo\\RRHH\\trabajadores.xlsx'

if not os.path.exists(file_path):
    print(f"Error: File not found at {file_path}")
    exit(1)

import datetime

try:
    wb = load_workbook(file_path, data_only=True)
    print(f"REPORT: Deep Analysis of {os.path.basename(file_path)}")
    print(f"Date: {datetime.datetime.now()}")
    print("-" * 60)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 4:
            print(f"\nSheet {sheet_name}: Too small to analyze (Rows: {len(rows)})")
            continue

        # Based on previous analysis, headers are likely in Row 3 (index 2)
        headers = [str(h).strip() if h else f"Col_{i}" for i, h in enumerate(rows[2])]
        data_rows = rows[3:]
        
        print(f"\nANALYSIS FOR SHEET: {sheet_name}")
        print(f"Total Rows (including headers): {len(rows)}")
        print(f"Data Rows: {len(data_rows)}")
        
        # 1. Column Filling Stats
        print("\n1. Column Population (Non-empty counts):")
        stats = {}
        for col_idx, h_name in enumerate(headers):
            non_empty = [r[col_idx] for r in data_rows if r[col_idx] is not None and str(r[col_idx]).strip() != ""]
            if non_empty:
                stats[col_idx] = {
                    'name': h_name,
                    'count': len(non_empty),
                    'unique': len(set(non_empty))
                }
                if len(non_empty) > 0:
                    print(f"  - {h_name:30} | Populated: {len(non_empty):3} | Unique: {len(set(non_empty)):3}")

        # 2. Categorical Analysis
        print("\n2. Categorical Data (Top Unique Values):")
        # Column 1: Farmacia, Column 9: Cargo, Column 10: Departamento
        cat_cols = [1, 9, 10]
        for c_idx in cat_cols:
            if c_idx < len(headers):
                vals = [str(r[c_idx]).strip() for r in data_rows if r[c_idx]]
                unique_vals = sorted(list(set(vals)))
                print(f"  * {headers[c_idx]}: {len(unique_vals)} unique values")
                for v in unique_vals[:10]: # Show first 10
                    print(f"    - {v}")
                if len(unique_vals) > 10: print("    ...")

        # 3. Data Quality / Anomalies
        print("\n3. Data Quality Issues:")
        # Check for missing Cedula (Col 6)
        missing_cid = [i+4 for i, r in enumerate(data_rows) if not r[6]]
        if missing_cid:
            print(f"  [!] Rows missing CID (Cedula): {missing_cid[:5]} (Total: {len(missing_cid)})")
        else:
            print("  [OK] All records have a CID (Cedula).")

        # Check for Date formats
        date_cols = [5, 8] # Birth, Hire
        for c_idx in date_cols:
            if c_idx < len(headers):
                invalid_dates = []
                for i, r in enumerate(data_rows):
                    val = r[c_idx]
                    if val and not isinstance(val, (datetime.date, datetime.datetime)):
                        invalid_dates.append(f"Row {i+4}: {val}")
                if invalid_dates:
                    print(f"  [!] {headers[c_idx]} has non-date objects in {len(invalid_dates)} rows.")
                    print(f"      Example: {invalid_dates[0]}")
                else:
                    print(f"  [OK] All populated {headers[c_idx]} values are valid date objects.")

        # 4. Composite Fields Detection
        # I noticed Col 26 (Cedula del hijo) sometimes has multiple values
        if len(headers) > 26:
            multi_vals = 0
            for r in data_rows:
                if r[26] and ("  " in str(r[26]) or "\n" in str(r[26])):
                    multi_vals += 1
            if multi_vals:
                print(f"  [!] Detected {multi_vals} rows with multiple values in 'CEDULA DEL HIJO' (needs split).")

except Exception as e:
    import traceback
    traceback.print_exc()





"""
Servicio de generación de reportes Excel para nómina.

Genera resúmenes de nómina por periodo con desglose por departamento, sede y concepto.
Usa openpyxl para la generación.
"""
from decimal import Decimal
from typing import Optional

from payroll_core.models import (
    PayrollPeriod, PayrollReceipt, PayrollReceiptLine,
    Branch, Department,
)


class PayrollExcelReport:
    """
    Genera reportes Excel de resumen de nómina.
    """
    
    @staticmethod
    def generate_period_summary(period_id: int) -> bytes:
        """
        Genera un reporte Excel con el resumen de un periodo de nómina.
        
        Hojas:
        1. Resumen General — totales por concepto
        2. Por Departamento — subtotales por departamento
        3. Por Empleado — detalle individual
        
        Returns:
            Bytes del archivo XLSX
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        except ImportError:
            raise ImportError("openpyxl no está instalado. Ejecute: pip install openpyxl")
        
        period = PayrollPeriod.objects.get(pk=period_id)
        receipts = PayrollReceipt.objects.filter(period=period).select_related(
            'employee', 'employee__department', 'employee__branch'
        ).order_by('employee__department__name', 'employee__last_name')
        
        wb = Workbook()
        
        # Estilos
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='1A237E', end_color='1A237E', fill_type='solid')
        subtotal_fill = PatternFill(start_color='E8EAF6', end_color='E8EAF6', fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )
        
        # =====================
        # HOJA 1: Detalle por Empleado
        # =====================
        ws1 = wb.active
        ws1.title = "Detalle por Empleado"
        
        headers = [
            'Cédula', 'Empleado', 'Departamento', 'Sede', 'Cargo',
            'Total Asignaciones', 'Total Deducciones', 'Neto a Pagar',
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        row = 2
        for receipt in receipts:
            emp = receipt.employee
            ws1.cell(row=row, column=1, value=emp.national_id)
            ws1.cell(row=row, column=2, value=emp.full_name)
            ws1.cell(row=row, column=3, value=str(emp.department) if emp.department else '')
            ws1.cell(row=row, column=4, value=str(emp.branch) if emp.branch else '')
            ws1.cell(row=row, column=5, value=emp.position or '')
            ws1.cell(row=row, column=6, value=float(receipt.total_income_ves or 0))
            ws1.cell(row=row, column=7, value=float(receipt.total_deductions_ves or 0))
            ws1.cell(row=row, column=8, value=float(receipt.net_pay_ves or 0))
            row += 1
        
        # Totales
        ws1.cell(row=row, column=5, value='TOTALES').font = Font(bold=True)
        ws1.cell(row=row, column=6, value=f'=SUM(F2:F{row-1})')
        ws1.cell(row=row, column=7, value=f'=SUM(G2:G{row-1})')
        ws1.cell(row=row, column=8, value=f'=SUM(H2:H{row-1})')
        
        # Ajustar ancho de columnas
        for col in range(1, len(headers) + 1):
            ws1.column_dimensions[chr(64 + col)].width = 18
        
        # =====================
        # HOJA 2: Resumen por Departamento
        # =====================
        ws2 = wb.create_sheet("Resumen por Departamento")
        
        dept_headers = ['Departamento', 'Empleados', 'Total Asignaciones', 'Total Deducciones', 'Neto Total']
        for col, header in enumerate(dept_headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Agregar datos agrupados
        dept_totals = {}
        for receipt in receipts:
            dept_name = str(receipt.employee.department) if receipt.employee.department else 'Sin Departamento'
            if dept_name not in dept_totals:
                dept_totals[dept_name] = {
                    'count': 0,
                    'earnings': Decimal('0'),
                    'deductions': Decimal('0'),
                    'net': Decimal('0'),
                }
            dept_totals[dept_name]['count'] += 1
            dept_totals[dept_name]['earnings'] += receipt.total_income_ves or 0
            dept_totals[dept_name]['deductions'] += receipt.total_deductions_ves or 0
            dept_totals[dept_name]['net'] += receipt.net_pay_ves or 0
        
        row = 2
        for dept, data in sorted(dept_totals.items()):
            ws2.cell(row=row, column=1, value=dept)
            ws2.cell(row=row, column=2, value=data['count'])
            ws2.cell(row=row, column=3, value=float(data['earnings']))
            ws2.cell(row=row, column=4, value=float(data['deductions']))
            ws2.cell(row=row, column=5, value=float(data['net']))
            row += 1
        
        for col in range(1, 6):
            ws2.column_dimensions[chr(64 + col)].width = 22
        
        # =====================
        # HOJA 3: Resumen por Concepto
        # =====================
        ws3 = wb.create_sheet("Resumen por Concepto")
        
        concept_headers = ['Código', 'Concepto', 'Tipo', 'Total', 'Empleados Afectados']
        for col, header in enumerate(concept_headers, 1):
            cell = ws3.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Agregar datos por concepto
        lines = PayrollReceiptLine.objects.filter(
            receipt__period=period
        ).values('concept_code', 'concept_name', 'kind').annotate(
            total=__import__('django.db.models', fromlist=['Sum']).Sum('amount_ves'),
            count=__import__('django.db.models', fromlist=['Count']).Count('id'),
        ).order_by('kind', 'concept_code')
        
        row = 2
        for line in lines:
            ws3.cell(row=row, column=1, value=line['concept_code'])
            ws3.cell(row=row, column=2, value=line['concept_name'])
            ws3.cell(row=row, column=3, value='Asignación' if line['kind'] == 'EARNING' else 'Deducción')
            ws3.cell(row=row, column=4, value=float(line['total'] or 0))
            ws3.cell(row=row, column=5, value=line['count'])
            row += 1
        
        for col in range(1, 6):
            ws3.column_dimensions[chr(64 + col)].width = 22
        
        # Guardar a bytes
        import io
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer.getvalue()

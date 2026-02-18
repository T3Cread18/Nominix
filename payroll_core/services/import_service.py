import pandas as pd
import numpy as np
import io
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.apps import apps
from django.db import models, transaction, IntegrityError
from django.core.exceptions import ValidationError, ObjectDoesNotExist
from django.utils.translation import gettext as _
from datetime import datetime
import logging

logger = logging.getLogger(__name__)

class ImportService:
    """
    Service to handle data import from files (Excel/CSV) into Django models.
    Inspired by Odoo's import functionality.
    """

    ALLOWED_MODELS = {
        'employee': ('payroll_core', 'Employee'),
        'branch': ('payroll_core', 'Branch'),
        'department': ('payroll_core', 'Department'),
        'jobposition': ('payroll_core', 'JobPosition'),
    }

    def get_model(self, model_key):
        if model_key not in self.ALLOWED_MODELS:
            raise ValueError(f"Model '{model_key}' is not allowed for import.")
        app_label, model_name = self.ALLOWED_MODELS[model_key]
        return apps.get_model(app_label, model_name)

    def get_model_fields(self, model_key):
        """
        Returns a list of importable fields for the given model.
        """
        model = self.get_model(model_key)
        fields = []
        
        for field in model._meta.get_fields():
            if field.concrete and not field.auto_created and not field.is_relation:
                fields.append({
                    'name': field.name,
                    'label': field.verbose_name.title() if hasattr(field, 'verbose_name') else field.name,
                    'type': field.get_internal_type(),
                    'required': not field.blank and not field.null and field.default == field.empty_strings_allowed,
                    'is_relation': False
                })
            elif field.concrete and field.is_relation and not field.auto_created:
                 # Foreign Keys (Simple handling for now, e.g., by ID or code)
                 fields.append({
                    'name': field.name,
                    'label': field.verbose_name.title() if hasattr(field, 'verbose_name') else field.name,
                    'type': 'ForeignKey',
                    'required': not field.blank and not field.null,
                    'related_model': field.related_model._meta.model_name,
                    'is_relation': True
                })
        
        return fields

    def _read_clean_file(self, file):
        """
        Reads the file into a DataFrame and sanitizes it for JSON/Django compatibility.
        - Replaces Infinity/NaN with None
        - Converts float columns to object to allow None
        """
        try:
            if hasattr(file, 'seek'):
                file.seek(0)

            if file.name.endswith('.csv'):
                df = pd.read_csv(file)
            elif file.name.endswith(('.xls', '.xlsx')):
                df = pd.read_excel(file)
            else:
                raise ValueError("Unsupported file format. Please upload CSV or Excel.")

            # Replace Infinity with NaN first
            df.replace([np.inf, -np.inf], np.nan, inplace=True)
            
            # Convert to object to allow None (instead of NaN) which is JSON compliant
            # This preserves Timestamps as objects but converts floats/ints to objects
            df = df.astype(object).where(pd.notnull(df), None)
            
            return df
        except Exception as e:
             # logging is good practice but let's just re-raise for now as per existing pattern
             raise e

    def preview_file(self, file):
        """
        Parses the file and returns headers and a preview of data.
        """
        try:
            df = self._read_clean_file(file)
            
            headers = [str(c) for c in df.columns]
            
            # For preview, we can just return the dict since None is handled
            preview_data = df.head(5).to_dict(orient='records')
            
            return {
                'headers': headers,
                'preview': preview_data,
                'total_rows': len(df)
            }
        except Exception as e:
            logger.error(f"Error previewing file: {str(e)}")
            raise

    def _resolve_foreign_keys(self, model, data):
        """
        Resolves foreign key fields in the data dict.
        Input data might contain ID (int), Code (str), or Name (str).
        We try to find the related instance.
        """
        resolved_data = data.copy()
        
        for field in model._meta.get_fields():
            if not field.is_relation or field.many_to_many or field.name not in data:
                continue
                
            value = data[field.name]
            if value is None:
                continue
                
            related_model = field.related_model
            
            # If value is already an instance, skip
            if isinstance(value, related_model):
                continue
                
            # Try 1: Is it a PK?
            if isinstance(value, int) or (isinstance(value, str) and value.isdigit()):
                try:
                    instance = related_model.objects.get(pk=value)
                    resolved_data[field.name] = instance
                    continue
                except related_model.DoesNotExist:
                    pass # Try other lookups

            # Try 2: Lookup by 'code' if exists
            if hasattr(related_model, 'code'):
                try:
                    instance = related_model.objects.get(code=value)
                    resolved_data[field.name] = instance
                    continue
                except related_model.DoesNotExist:
                    pass
            
            # Try 3: Lookup by 'name' if exists
            # Note: Name might not be unique, so we might want to filter and take first?
            # Or strict unique? Let's try strict first, then lenient.
            if hasattr(related_model, 'name'):
                try:
                    instance = related_model.objects.get(name=value)
                    resolved_data[field.name] = instance
                    continue
                except related_model.MultipleObjectsReturned:
                     # If duplicate names, this is ambiguous.
                     raise ValueError(f"Multiple {related_model._meta.verbose_name} found with name '{value}'. Please use Code or ID.")
                except related_model.DoesNotExist:
                    pass
            
            # Try 4: Lookup by specific fields for certain models
            if related_model.__name__ == 'Employee':
                 # By National ID
                try:
                    instance = related_model.objects.get(national_id=value)
                    resolved_data[field.name] = instance
                    continue
                except related_model.DoesNotExist:
                    pass
            
            # If we reached here, we couldn't resolve it. 
            # We assume it might be a valid raw ID that exists but we failed to fetch? 
            # Or we simply leave it and let django validation fail.
            # But normally we want to replace the raw value with the Instance if found.
            # If not found, better to fail hard here or let standard validation catch "Must be instance".
            
            # Let's check cache/optimize later. For now, simplistic.
            pass
            
        return resolved_data

    def _clean_field_values(self, model, data):
        """
        Cleans field values based on strict model constraints.
        Mainly converts None to "" for CharField/TextField that are not nullable.
        """
        cleaned_data = data.copy()
        
        for field in model._meta.get_fields():
            if field.name not in data:
                continue
                
            value = data[field.name]
            
            # Handle CharField/TextField with null=False, blank=True
            if isinstance(field, (models.CharField, models.TextField)):
                if value is None and not field.null:
                    cleaned_data[field.name] = ""
            
            # Additional cleaners can go here
            
        return cleaned_data

    def validate_import(self, file, mapping, model_key):
        """
        Dry-run validation of the import.
        mapping: {'csv_column': 'model_field', ...}
        """
        model = self.get_model(model_key)
        errors = []
        
        try:
            df = self._read_clean_file(file)
            
            for index, row in df.iterrows():
                row_errors = []
                data = {}
                
                # Build data dict from mapping
                for csv_col, model_field in mapping.items():
                    if not model_field: continue
                    val = row.get(csv_col)
                    if val == "": val = None
                    data[model_field] = val

                # Validate using Model instance
                try:
                    # Resolve FKs
                    resolved_data = self._resolve_foreign_keys(model, data)
                    
                    # Clean strict values (None vs "")
                    cleaned_data = self._clean_field_values(model, resolved_data)
                    
                    instance = model(**cleaned_data)
                    instance.full_clean()
                except ValueError as e:
                     row_errors.append(str(e))
                except ValidationError as e:
                    for field, err_list in e.message_dict.items():
                        for err in err_list:
                             row_errors.append(f"Field '{field}': {err}")
                except Exception as e:
                    row_errors.append(f"General Error: {str(e)}")

                if row_errors:
                    errors.append({
                        'row': index + 2,
                        'errors': row_errors,
                        'data': row.to_dict()
                    })
                    
            return {'valid': len(errors) == 0, 'errors': errors}

        except Exception as e:
            logger.error(f"Error validating import: {str(e)}")
            raise e

    def execute_import(self, file, mapping, model_key):
        """
        Executes the import with transaction safety.
        """
        model = self.get_model(model_key)
        created_count = 0
        updated_count = 0
        errors = []
        
        try:
           df = self._read_clean_file(file)
        except Exception as e:
            raise e

        with transaction.atomic():
            for index, row in df.iterrows():
                try:
                    data = {}
                    for csv_col, model_field in mapping.items():
                        if not model_field: continue
                        val = row.get(csv_col)
                        if val == "": val = None
                        data[model_field] = val
                    
                    # Resolve FKs
                    resolved_data = self._resolve_foreign_keys(model, data)
                    
                    # Clean strict values (None vs "")
                    cleaned_data = self._clean_field_values(model, resolved_data)

                    # Naive create for now. 
                    # TODO: Implement "Update if exists" logic based on a key (e.g. ID or Code)
                    model.objects.create(**cleaned_data)
                    created_count += 1
                    
                except Exception as e:
                    errors.append(f"Row {index+2}: {str(e)}")
                    # For now: strict all-or-nothing.
                    raise e
                    
        return {
            'success': True,
            'created': created_count,
            'updated': updated_count
        }

    def generate_template(self, model_key, file_format='xlsx'):
        """
        Generates a downloadable template file for the given model.
        Includes headers, field type info, and an example row.
        """
        fields = self.get_model_fields(model_key)
        model = self.get_model(model_key)
        
        # Friendly model names
        model_labels = {
            'employee': 'Empleados',
            'branch': 'Sedes',
            'department': 'Departamentos',
            'jobposition': 'Cargos',
        }
        
        # Example data per model
        example_data = {
            'employee': {
                'employee_code': 'EMP-001',
                'first_name': 'Juan',
                'last_name': 'Pérez',
                'id_number': 'V-12345678',
                'gender': 'M',
                'marital_status': 'S',
                'address': 'Av. Principal #123',
                'phone': '0412-1234567',
                'email': 'juan.perez@example.com',
                'hire_date': '2024-01-15',
                'base_salary': '500.00',
                'status': 'active',
            },
            'branch': {
                'name': 'Sede Principal',
                'code': 'SP-001',
                'address': 'Av. Bolívar #100',
                'phone': '0212-1234567',
            },
            'department': {
                'name': 'Recursos Humanos',
                'code': 'RRHH',
            },
            'jobposition': {
                'name': 'Analista de Nómina',
                'code': 'AN-001',
            },
        }
        
        examples = example_data.get(model_key, {})
        
        if file_format == 'csv':
            # Simple CSV
            output = io.StringIO()
            headers = [f['name'] for f in fields]
            output.write(','.join(headers) + '\n')
            
            # Example row
            row = [str(examples.get(f['name'], '')) for f in fields]
            output.write(','.join(row) + '\n')
            
            return output.getvalue().encode('utf-8'), f'plantilla_{model_key}.csv', 'text/csv'
        
        else:
            # Excel with formatting
            output = io.BytesIO()
            
            headers = [f['name'] for f in fields]
            labels = [f['label'] for f in fields]
            types = []
            for f in fields:
                t = f['type']
                req = '(Requerido)' if f.get('required') else '(Opcional)'
                if f.get('is_relation'):
                    t = f"FK → {f['related_model']}"
                types.append(f"{t} {req}")
            
            example_row = [examples.get(f['name'], '') for f in fields]
            
            df = pd.DataFrame([example_row], columns=headers)
            
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Datos', startrow=2)
                
                ws = writer.sheets['Datos']
                
                # Styles
                header_font = Font(bold=True, color='FFFFFF', size=11)
                header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
                info_font = Font(italic=True, color='666666', size=9)
                label_font = Font(bold=True, color='1F4E79', size=10)
                thin_border = Border(
                    left=Side(style='thin', color='D9D9D9'),
                    right=Side(style='thin', color='D9D9D9'),
                    top=Side(style='thin', color='D9D9D9'),
                    bottom=Side(style='thin', color='D9D9D9'),
                )
                
                # Row 1: Labels (friendly names)
                for col_idx, label in enumerate(labels, 1):
                    cell = ws.cell(row=1, column=col_idx, value=label)
                    cell.font = label_font
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = thin_border
                
                # Row 2: Type info
                for col_idx, type_info in enumerate(types, 1):
                    cell = ws.cell(row=2, column=col_idx, value=type_info)
                    cell.font = info_font
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = thin_border
                
                # Row 3: Headers (field names) - already written by pandas
                for col_idx in range(1, len(headers) + 1):
                    cell = ws.cell(row=3, column=col_idx)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = thin_border
                
                # Row 4: Example data - style it
                example_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
                for col_idx in range(1, len(headers) + 1):
                    cell = ws.cell(row=4, column=col_idx)
                    cell.fill = example_fill
                    cell.border = thin_border
                
                # Auto-width columns
                for col_idx, header in enumerate(headers, 1):
                    max_len = max(len(str(header)), len(str(labels[col_idx-1])), len(str(types[col_idx-1])), 15)
                    ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else 'A'].width = max_len + 4
            
            output.seek(0)
            model_label = model_labels.get(model_key, model_key)
            filename = f'plantilla_{model_key}.xlsx'
            content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            
            return output.getvalue(), filename, content_type

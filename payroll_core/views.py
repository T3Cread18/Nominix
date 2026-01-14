from django.db import transaction
from rest_framework import views, viewsets, response, status, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from decimal import Decimal
from django_filters.rest_framework import DjangoFilterBackend
from django.template.loader import render_to_string
from django.http import HttpResponse
from django.db.models import ProtectedError
import weasyprint
import csv
from .services import BCVRateService
from rest_framework.views import APIView
from .models import (
    Employee, LaborContract, ExchangeRate, PayrollConcept, 
    EmployeeConcept, Branch, Currency, PayrollPeriod, Payslip,
    PayrollNovelty, Company, Department, Loan, LoanPayment, JobPosition
)
from .serializers import (
    EmployeeSerializer, BranchSerializer, LaborContractSerializer,
    CurrencySerializer, PayrollConceptSerializer, EmployeeConceptSerializer,
    PayrollPeriodSerializer, PayslipSerializer, PayrollNoveltySerializer, CompanySerializer,
    DepartmentSerializer, LoanSerializer, LoanPaymentSerializer, JobPositionSerializer
)
from .engine import PayrollEngine

class CurrencyViewSet(viewsets.ModelViewSet):
    queryset = Currency.objects.all()
    serializer_class = CurrencySerializer

class PayrollConceptViewSet(viewsets.ModelViewSet):
    queryset = PayrollConcept.objects.all()
    serializer_class = PayrollConceptSerializer
    filterset_fields = ['kind', 'active']

    def perform_destroy(self, instance):
        if instance.is_system:
            raise serializers.ValidationError(
                "No se puede eliminar un concepto de sistema."
            )
        super().perform_destroy(instance)

class EmployeeConceptViewSet(viewsets.ModelViewSet):
    queryset = EmployeeConcept.objects.all()
    serializer_class = EmployeeConceptSerializer
    filterset_fields = ['employee', 'is_active']

class BranchViewSet(viewsets.ModelViewSet):
    queryset = Branch.objects.all()
    serializer_class = BranchSerializer
class LatestExchangeRateView(APIView):
    """
    Devuelve la tasa de cambio más reciente para una moneda dada.
    Si no existe tasa para el día de hoy, intenta obtenerla del BCV.
    """
    def get(self, request):
        currency_code = request.query_params.get('currency', 'USD')
        today = timezone.now().date()

        # 1. Intentar buscar en DB la tasa de HOY
        rate_obj = ExchangeRate.objects.filter(
            currency__code=currency_code,
            date_valid__date=today
        ).first()

        # 2. Si no existe la de hoy, llamamos al Servicio BCV (Scraping)
        if not rate_obj:
            print(f"🔄 Tasa {currency_code} desactualizada. Consultando BCV...")
            try:
                # Tu servicio guarda en DB, así que solo lo ejecutamos
                results = BCVRateService.fetch_and_update_rates()
                
                # Verificamos si la operación fue exitosa para la moneda solicitada
                if results.get(currency_code, {}).get('status') == 'success':
                    # Volvemos a consultar la DB recién actualizada
                    rate_obj = ExchangeRate.objects.filter(
                        currency__code=currency_code
                    ).order_by('-date_valid').first()
            except Exception as e:
                print(f"Error consultando servicio BCV: {e}")

        # 3. Si aún no hay tasa (ej. BCV caído y DB vacía), buscar la última histórica
        if not rate_obj:
             rate_obj = ExchangeRate.objects.filter(
                currency__code=currency_code
            ).order_by('-date_valid').first()

        # 4. Respuesta final
        if rate_obj:
            return Response({
                "currency": currency_code,
                "rate": rate_obj.rate,
                "date": rate_obj.date_valid,
                "source": rate_obj.source
            })
        else:
            # Fallback extremo si no hay nada en DB ni internet
            return Response(
                {"currency": currency_code, "rate": 60.00, "note": "Valor por defecto"}, 
                status=status.HTTP_200_OK
            )
class LaborContractViewSet(viewsets.ModelViewSet):
    queryset = LaborContract.objects.all()
    serializer_class = LaborContractSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['employee', 'is_active', 'branch']

class DepartmentViewSet(viewsets.ModelViewSet):
    queryset = Department.objects.all()
    serializer_class = DepartmentSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    search_fields = ['name']
    search_fields = ['name']
    filterset_fields = ['branch']

class JobPositionViewSet(viewsets.ModelViewSet):
    queryset = JobPosition.objects.all()
    serializer_class = JobPositionSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    search_fields = ['name', 'code']
    filterset_fields = ['department']


class EmployeeViewSet(viewsets.ModelViewSet):
    queryset = Employee.objects.all()
    serializer_class = EmployeeSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    search_fields = ['first_name', 'last_name', 'national_id']
    filterset_fields = ['is_active', 'branch']
    @action(detail=True, methods=['get', 'post'], url_path='simulate-payslip')
    def simulate_payslip(self, request, pk=None):
        """
        GET/POST /api/employees/{id}/simulate-payslip/
        Calcula la nómina proyectada usando el Engine Orquestador.
        """
        employee = self.get_object()
        # 1. Obtener variables de entrada (POST json o GET params)
        # 1. Obtener variables de entrada (POST json o GET params)
        input_variables = None
        period_id = None
        
        if request.method == 'POST':
            data = request.data.copy()
            period_id = data.pop('period_id', None)
            if data:
                input_variables = data
        else:
            # Convertir query params a dict
            data = {k: v for k, v in request.query_params.items()}
            period_id = data.pop('period_id', None)
            if data:
                input_variables = data

        # 2. Buscar Contrato Activo
        contract = LaborContract.objects.filter(employee=employee, is_active=True).first()
        if not contract:
            return Response(
                {"error": "El empleado no tiene contrato activo."}, 
                status=status.HTTP_400_BAD_REQUEST
            )
            
        # 2.1 Buscar Periodo (si se indica)
        period = None
        if period_id:
            try:
                period = PayrollPeriod.objects.get(pk=period_id)
            except PayrollPeriod.DoesNotExist:
                pass

        try:
            # 3. Inicializar Motor
            # Si input_variables es None, el motor cargará las novedades de DB (si hay periodo)
            payment_date = period.payment_date if period else timezone.now().date()
            
            engine = PayrollEngine(
                contract=contract, 
                period=period,
                payment_date=payment_date, 
                input_variables=input_variables
            )

            # 4. Delegar el cálculo al Engine (Orquestación)
            # Esto automáticamente inyecta Sueldo Base, IVSS, FAOV y Novedades
            result = engine.calculate_payroll()

            # 5. Forzar conversión de Decimals a Float para facilitar el consumo en JS formatCurrency
            result['totals'] = {k: float(v) if isinstance(v, Decimal) else v for k, v in result.get('totals', {}).items()}
            for line in result.get('lines', []):
                for k, v in line.items():
                    if isinstance(v, Decimal):
                        line[k] = float(v)

            # 5. Retornar el resultado directamente (Ya viene formateado en JSON)
            return Response(result)

        except Exception as e:
            # Log de error para debug en consola
            print(f"Error calculando simulación: {e}")
            return Response(
                {"error": f"Error interno de cálculo: {str(e)}"}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def update(self, request, *args, **kwargs):
        partial = True # Forzamos parcial para que no borre datos no enviados
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        return Response(serializer.data)
    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        try:
            self.perform_destroy(instance)
            return Response(status=status.HTTP_204_NO_CONTENT)
        except ProtectedError:
            # Capturamos el error de integridad y devolvemos un mensaje amigable
            return Response(
                {"error": "No se puede eliminar este colaborador porque tiene histórico de nóminas o contratos. Proceda a desactivarlo (Dar de baja)."},
                status=status.HTTP_400_BAD_REQUEST
            )

class PayrollPeriodViewSet(viewsets.ModelViewSet):
    queryset = PayrollPeriod.objects.all()
    serializer_class = PayrollPeriodSerializer

    @action(detail=True, methods=['get', 'post'], url_path='preview-payroll')
    def preview_payroll(self, request, pk=None):
        """
        GET/POST /api/payroll-periods/{id}/preview-payroll/
        Calcula la nómina proyectada para todos los empleados del periodo.
        """
        try:
            from .services.payroll import PayrollProcessor
            manual_rate = request.data.get('manual_rate') if request.method == 'POST' else request.query_params.get('manual_rate')
            result = PayrollProcessor.preview_period(pk, manual_rate=manual_rate)
            return Response(result, status=status.HTTP_200_OK)
        except ValueError as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"error": f"Falla inesperada en previsualización: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['post'], url_path='close-period')
    def close_period(self, request, pk=None):
        """
        POST /api/payroll-periods/{id}/close-period/
        Inicia el proceso de cálculo masivo y cierre inmutable.
        """
        try:
            from .services import PayrollProcessor
            manual_rate = request.data.get('manual_rate')
            result = PayrollProcessor.process_period(pk, user=request.user, manual_rate=manual_rate)
            return Response(result, status=status.HTTP_200_OK)
        except ValueError as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"error": f"Falla inesperada en cierre: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    @action(detail=True, methods=['get'], url_path='export-pdf')
    def export_payslips_pdf(self, request, pk=None):
        """
        Genera un PDF masivo con todos los recibos del periodo.
        Params: ?tipo=todos|salario|complemento|cestaticket
        """
        period = self.get_object()
        tipo_recibo = request.query_params.get('tipo', 'todos')
        
        # Obtener recibos asociados a este periodo
        payslips_qs = period.payslips.select_related('employee', 'employee__department').prefetch_related('details').all()
        
        if not payslips_qs.exists():
            return Response({"error": "No hay recibos generados para este periodo."}, status=404)

        # 1. Obtener conceptos configurados para el recibo (Estructura Fija)
        configured_concepts = PayrollConcept.objects.filter(
            appears_on_receipt=True
        ).order_by('receipt_order')
        
        # Mapeo de códigos configurados para referencia rápida
        configured_codes = {c.code: c for c in configured_concepts}

        # 2. Pre-procesar cada payslip
        processed_payslips = []
        for payslip in payslips_qs:
            actual_details = {d.concept_code: d for d in payslip.details.all()}
            used_codes = set()
            payslip_rows = []
            
            # Primero, agregar conceptos configurados (Orden Fijo)
            for concept in configured_concepts:
                detail = actual_details.get(concept.code)
                used_codes.add(concept.code)
                
                if detail:
                    row = {
                        'name': detail.concept_name,
                        'code': detail.concept_code,
                        'kind': detail.kind,
                        'quantity': getattr(detail, 'quantity', 0) or 0,
                        'unit': getattr(detail, 'unit', '') or '',
                        'amount_ves': detail.amount_ves,
                        'tipo_recibo': detail.tipo_recibo,
                    }
                elif concept.show_even_if_zero:
                    row = {
                        'name': concept.name,
                        'code': concept.code,
                        'kind': concept.kind,
                        'quantity': 0,
                        'unit': '',
                        'amount_ves': Decimal('0.00'),
                        'tipo_recibo': 'salario',
                    }
                else:
                    # Si no hay detalle y no se fuerza el cero, no agregar
                    used_codes.remove(concept.code) # Marcar como no usado realmente
                    continue
                
                row['earning_amount'] = row['amount_ves'] if row['kind'] == 'EARNING' else None
                row['deduction_amount'] = row['amount_ves'] if row['kind'] == 'DEDUCTION' else None
                payslip_rows.append(row)
            
            # Segundo, agregar detalles calculados que NO estaban en la configuración fija (Fallback)
            for code, detail in actual_details.items():
                if code not in used_codes:
                    row = {
                        'name': detail.concept_name,
                        'code': detail.concept_code,
                        'kind': detail.kind,
                        'quantity': getattr(detail, 'quantity', 0) or 0,
                        'unit': getattr(detail, 'unit', '') or '',
                        'amount_ves': detail.amount_ves,
                        'tipo_recibo': detail.tipo_recibo,
                        'earning_amount': detail.amount_ves if detail.kind == 'EARNING' else None,
                        'deduction_amount': detail.amount_ves if detail.kind == 'DEDUCTION' else None,
                    }
                    payslip_rows.append(row)
            
            payslip.processed_rows = payslip_rows
            processed_payslips.append(payslip)

        # Contexto para el Template HTML
        company_config = Company.objects.first()
        context = {
            'payslips': processed_payslips,
            'period_name': period.name,
            'start_date': period.start_date,
            'end_date': period.end_date,
            'payment_date': period.payment_date,
            'tenant_name': request.tenant.name if hasattr(request, 'tenant') else "Nóminix",
            'tenant_rif': request.tenant.rif if hasattr(request, 'tenant') else "",
            'tenant_address': request.tenant.address if hasattr(request, 'tenant') else "",
            'company_config': company_config,
            'tipo_recibo': tipo_recibo,
        }

        # Seleccionar template según tipo de recibo
        template_map = {
            'todos': 'payroll/payslip_batch.html',
            'salario': 'payroll/recibo_salario.html',
            'complemento': 'payroll/recibo_complemento.html',
            'cestaticket': 'payroll/recibo_cestaticket.html',
        }
        template_name = template_map.get(tipo_recibo, 'payroll/payslip_batch.html')

        # 1. Renderizar HTML string
        html_string = render_to_string(template_name, context)

        # 2. Generar PDF con WeasyPrint
        pdf_file = weasyprint.HTML(string=html_string).write_pdf()

        # 3. Retornar respuesta HTTP con Content-Type correcto
        suffix = f'_{tipo_recibo}' if tipo_recibo != 'todos' else ''
        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="nomina{suffix}_{period.id}.pdf"'
        return response

    @action(detail=True, methods=['get'], url_path='export-finance')
    def export_finance_report(self, request, pk=None):
        """
        Genera un reporte Excel (.xlsx) nativo para Finanzas.
        Permite definir tipos de celda explícitos.
        """
        period = self.get_object()
        # Usamos select_related para optimizar la DB
        payslips = period.payslips.select_related('employee').all()

        if not payslips.exists():
            return Response({"error": "No hay recibos generados para este periodo."}, status=404)

        # 1. Configuración del archivo Excel
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="finanzas_nomina_{period.id}.xlsx"'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte Finanzas"

        # 2. Encabezados
        headers = ['Cédula', 'Empleado', 'Sede', 'Banco', 'Tipo Cuenta', 'Número de Cuenta', 'Monto VES', 'Tasa BCV', 'Monto USD Ref']
        ws.append(headers)

        # Estilo para encabezados (Negrita)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # 3. Iterar datos
        for p in payslips:
            emp = p.employee
            
            # Cálculos previos seguros
            monto_ves = p.net_pay_ves or Decimal(0)
            tasa = p.exchange_rate_applied or Decimal(0)
            monto_usd = (monto_ves / tasa).quantize(Decimal('0.01')) if tasa > 0 else Decimal(0)

            # Preparamos la fila
            row = [
                emp.national_id,
                emp.full_name,
                emp.branch.name if emp.branch else 'N/D',
                emp.bank_name or 'N/D',
                emp.bank_account_type or 'N/D',
                emp.bank_account_number or 'N/D', # Se inserta como string inicialmente
                monto_ves,
                tasa,
                monto_usd
            ]
            ws.append(row)
            
            # 4. DEFINICIÓN DE FORMATOS POR CELDA (La magia ocurre aquí)
            current_row = ws.max_row
            
            # Columna E (Número de Cuenta): Forzar formato Texto para evitar notación científica
            cell_acc = ws.cell(row=current_row, column=5)
            cell_acc.number_format = '@'  # '@' significa Texto en Excel
            
            # Columna F (Monto VES): Formato numérico con separadores
            cell_ves = ws.cell(row=current_row, column=6)
            cell_ves.number_format = '#,##0.00' 

            # Columna G (Tasa): Formato con más decimales si es necesario
            cell_rate = ws.cell(row=current_row, column=7)
            cell_rate.number_format = '#,##0.0000'

            # Columna H (USD): Formato numérico
            cell_usd = ws.cell(row=current_row, column=8)
            cell_usd.number_format = '#,##0.00'

        # 5. Ajustar ancho de columnas automáticamente (Opcional pero útil)
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # 6. Guardar en la respuesta
        wb.save(response)
        return response

class PayslipReadOnlyViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet solo lectura para consultar recibos históricos.
    """
    queryset = Payslip.objects.all()
    serializer_class = PayslipSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['period', 'employee', 'currency_code']

    @action(detail=True, methods=['get'], url_path='export-pdf')
    def export_pdf(self, request, pk=None):
        """
        Genera el PDF individual para un recibo de pago.
        """
        import weasyprint
        from django.template.loader import render_to_string
        from django.http import HttpResponse

        payslip = self.get_object()
        
        # 1. Obtener conceptos configurados
        configured_concepts = PayrollConcept.objects.filter(
            appears_on_receipt=True
        ).order_by('receipt_order')

        # 2. Pre-procesar filas
        actual_details = {d.concept_code: d for d in payslip.details.all()}
        used_codes = set()
        payslip_rows = []
        
        # Primero, conceptos configurados
        for concept in configured_concepts:
            detail = actual_details.get(concept.code)
            used_codes.add(concept.code)
            
            if detail:
                row = {
                    'name': detail.concept_name,
                    'code': detail.concept_code,
                    'kind': detail.kind,
                    'quantity': getattr(detail, 'quantity', 0) or 0,
                    'unit': getattr(detail, 'unit', '') or '',
                    'amount_ves': detail.amount_ves,
                    'tipo_recibo': detail.tipo_recibo,
                }
            elif concept.show_even_if_zero:
                row = {
                    'name': concept.name,
                    'code': concept.code,
                    'kind': concept.kind,
                    'quantity': 0,
                    'unit': '',
                    'amount_ves': Decimal('0.00'),
                    'tipo_recibo': 'salario',
                }
            else:
                used_codes.remove(concept.code)
                continue
            
            row['earning_amount'] = row['amount_ves'] if row['kind'] == 'EARNING' else None
            row['deduction_amount'] = row['amount_ves'] if row['kind'] == 'DEDUCTION' else None
            payslip_rows.append(row)

        # Segundo, fallback para detalles no configurados
        for code, detail in actual_details.items():
            if code not in used_codes:
                row = {
                    'name': detail.concept_name,
                    'code': detail.concept_code,
                    'kind': detail.kind,
                    'quantity': getattr(detail, 'quantity', 0) or 0,
                    'unit': getattr(detail, 'unit', '') or '',
                    'amount_ves': detail.amount_ves,
                    'tipo_recibo': detail.tipo_recibo,
                    'earning_amount': detail.amount_ves if detail.kind == 'EARNING' else None,
                    'deduction_amount': detail.amount_ves if detail.kind == 'DEDUCTION' else None,
                }
                payslip_rows.append(row)
        
        payslip.processed_rows = payslip_rows
        company_config = Company.objects.first()
        
        context = {
            'payslips': [payslip],
            'period_name': payslip.period.name,
            'start_date': payslip.period.start_date,
            'end_date': payslip.period.end_date,
            'payment_date': payslip.period.payment_date,
            'tenant_name': request.tenant.name if hasattr(request, 'tenant') else "Nóminix",
            'tenant_rif': request.tenant.rif if hasattr(request, 'tenant') else "",
            'tenant_address': request.tenant.address if hasattr(request, 'tenant') else "",
            'company_config': company_config,
        }

        html_string = render_to_string('payroll/payslip_batch.html', context)
        pdf_file = weasyprint.HTML(string=html_string).write_pdf()

        response = HttpResponse(pdf_file, content_type='application/pdf')
        filename = f"recibo_{payslip.employee.national_id}_{payslip.period.id}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class PayrollNoveltyViewSet(viewsets.ModelViewSet):
    """
    Gestión de incidencias de nómina (Novedades).
    """
    queryset = PayrollNovelty.objects.all()
    serializer_class = PayrollNoveltySerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['employee', 'period', 'concept_code']

    @action(detail=False, methods=['post'], url_path='batch')
    def batch(self, request):
        """
        POST /api/payroll-novelties/batch/
        Carga masiva de novedades.
        Recibe: [{"employee_id": 1, "period_id": 5, "concept_code": "H_EXTRA", "amount": 4}, ...]
        """
        data = request.data
        if not isinstance(data, list):
            return Response(
                {"error": "Se esperaba una lista de novedades (array)."}, 
                status=status.HTTP_400_BAD_REQUEST
            )
            
        processed_ids = []
        try:
            with transaction.atomic():
                for item in data:
                    # Usamos update_or_create para manejar ediciones rápidas en el grid
                    obj, created = PayrollNovelty.objects.update_or_create(
                        employee_id=item.get('employee_id'),
                        period_id=item.get('period_id'),
                        concept_code=item.get('concept_code'),
                        defaults={'amount': Decimal(str(item.get('amount', 0)))}
                    )
                    processed_ids.append(obj.id)
            
            return Response({
                "status": "success", 
                "processed_count": len(processed_ids)
            }, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            return Response(
                {"error": f"Falla en procesamiento batch: {str(e)}"}, 
                status=status.HTTP_400_BAD_REQUEST
            )
class CompanyConfigView(views.APIView):
    """
    Endpoint único para obtener/editar la configuración de la empresa.
    Siempre devuelve el primer objeto encontrado o crea uno por defecto.
    """
    def get(self, request):
        company = Company.objects.first()
        if not company:
            # Crear datos default si no existen
            company = Company.objects.create(name="Mi Empresa C.A.", rif="J-00000000-0")
        serializer = CompanySerializer(company)
        return response.Response(serializer.data)

    def put(self, request):
        company = Company.objects.first()
        serializer = CompanySerializer(company, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return response.Response(serializer.data)
        return response.Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class LoanViewSet(viewsets.ModelViewSet):
    queryset = Loan.objects.all()
    serializer_class = LoanSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['employee', 'status', 'currency', 'frequency']
    search_fields = ['employee__first_name', 'employee__last_name', 'employee__national_id', 'description']

class LoanPaymentViewSet(viewsets.ModelViewSet):
    queryset = LoanPayment.objects.all()
    serializer_class = LoanPaymentSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['loan', 'payslip']

class PayrollVariablesView(APIView):
    """
    Devuelve el inventario de variables disponibles para fórmulas.
    """
    def get(self, request):
        return Response(PayrollEngine.get_variable_inventory())

class ValidateFormulaView(APIView):
    """
    Valida una fórmula y devuelve el resultado y la traza.
    """
    def post(self, request):
        formula = request.data.get('formula')
        context = request.data.get('context', {})
        if not formula:
            return Response({"error": "Fórmula no proporcionada"}, status=status.HTTP_400_BAD_REQUEST)
        
        result = PayrollEngine.validate_formula(formula, context)
        return Response(result)

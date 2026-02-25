from django.db import transaction
from rest_framework import views, viewsets, response, status, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from decimal import Decimal
from django_filters.rest_framework import DjangoFilterBackend
from django.template.loader import render_to_string
from django.http import HttpResponse
from django.db.models import ProtectedError
import weasyprint
import csv
from ..services import BCVRateService
from rest_framework.views import APIView
from ..models import (
    PayrollPeriod, PayrollReceipt, PayrollNovelty, Employee, 
    LaborContract, PayrollConcept, Company, Loan, Branch,
    ExchangeRate, EmployeeConcept, Currency, Department, LoanPayment, JobPosition,
    PayrollPolicy,
    # Social Benefits
    SocialBenefitsLedger, SocialBenefitsSettlement, InterestRateBCV
)
from ..serializers import (
    PayrollPeriodSerializer, PayrollReceiptSerializer, PayrollNoveltySerializer,
    EmployeeSerializer, LaborContractSerializer, PayrollConceptSerializer,
    CompanySerializer, LoanSerializer, BranchSerializer,
    CurrencySerializer, EmployeeConceptSerializer, DepartmentSerializer, 
    LoanPaymentSerializer, JobPositionSerializer, PayrollPolicySerializer,
    ACCUMULATOR_LABELS, BEHAVIOR_REQUIRED_PARAMS, ExchangeRateSerializer,
    # Social Benefits Serializers
    SocialBenefitsLedgerSerializer, SocialBenefitsSettlementSerializer,
    InterestRateBCVSerializer, AdvanceRequestSerializer, QuarterlyGuaranteeSerializer
)
from ..engine import PayrollEngine

# Import vacations
from vacations.models import VacationPayment, VacationRequest

# Import new views
from .import_views import *
from .endowment_views import EndowmentEventViewSet

class CurrencyViewSet(viewsets.ModelViewSet):
    queryset = Currency.objects.all()
    serializer_class = CurrencySerializer

class PayrollReceiptViewSet(viewsets.ModelViewSet):
    queryset = PayrollReceipt.objects.all()
    serializer_class = PayrollReceiptSerializer
    
    def get_queryset(self):
        queryset = PayrollReceipt.objects.all()
        period_id = self.request.query_params.get('period', None)
        if period_id:
            queryset = queryset.filter(period_id=period_id)
        return queryset

class PayrollConceptViewSet(viewsets.ModelViewSet):
    queryset = PayrollConcept.objects.all()
    serializer_class = PayrollConceptSerializer
    filterset_fields = ['kind', 'active']

    def perform_destroy(self, instance):
        if instance.is_system:
            from rest_framework import serializers
            raise serializers.ValidationError(
                "No se puede eliminar un concepto de sistema."
            )
        super().perform_destroy(instance)

class EmployeeConceptViewSet(viewsets.ModelViewSet):
    queryset = EmployeeConcept.objects.all()
    serializer_class = EmployeeConceptSerializer
    filterset_fields = ['employee', 'is_active']

class BranchViewSet(viewsets.ModelViewSet):
    queryset = Branch.objects.all()
    serializer_class = BranchSerializer


class ConceptConfigMetadataView(APIView):
    """
    Devuelve la metadata necesaria para el constructor de conceptos en el frontend.
    GET /api/concepts/config-metadata/
    """
    def get(self, request):
        # Convertir behaviors del enum a lista de opciones para UI
        behaviors = [
            {'value': choice[0], 'label': choice[1]}
            for choice in PayrollConcept.ConceptBehavior.choices
        ]
        
        # Convertir kinds del enum a lista de opciones
        kinds = [
            {'value': choice[0], 'label': choice[1]}
            for choice in PayrollConcept.ConceptKind.choices
        ]
        
        # Convertir computation methods
        computation_methods = [
            {'value': choice[0], 'label': choice[1]}
            for choice in PayrollConcept.ComputationMethod.choices
        ]
        
        # Convertir accumulator labels a lista para checkboxes
        accumulators = [
            {'code': code, 'label': label}
            for code, label in ACCUMULATOR_LABELS.items()
        ]
        
        # Parámetros requeridos por behavior
        behavior_params = BEHAVIOR_REQUIRED_PARAMS

        # Convertir calculation base methods
        calculation_bases = [
            {'value': choice[0], 'label': choice[1]}
            for choice in PayrollConcept.CalculationBase.choices
        ]

        return Response({
            'behaviors': behaviors,
            'kinds': kinds,
            'computation_methods': computation_methods,
            'accumulators': accumulators,
            'behavior_required_params': behavior_params,
            'calculation_base_options': calculation_bases,
        })


class PayrollPolicyView(APIView):
    """
    Vista para leer/editar la política de nómina de la empresa.
    GET/PUT /api/company/policies/
    """
    def get_policy(self, request):
        """Obtiene o crea la política para la empresa actual."""
        company = Company.objects.first()
        if not company:
            return None, Response(
                {'error': 'No hay empresa configurada.'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        policy, _ = PayrollPolicy.objects.get_or_create(company=company)
        return policy, None
    
    def get(self, request):
        policy, error_response = self.get_policy(request)
        if error_response:
            return error_response
        
        serializer = PayrollPolicySerializer(policy)
        return Response(serializer.data)
    
    def put(self, request):
        policy, error_response = self.get_policy(request)
        if error_response:
            return error_response
        
        serializer = PayrollPolicySerializer(policy, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class ExchangeRateViewSet(viewsets.ModelViewSet):
    queryset = ExchangeRate.objects.all().order_by('-date_valid')
    serializer_class = ExchangeRateSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['currency', 'source']
    ordering_fields = ['date_valid', 'rate']
    @action(detail=False, methods=['post'], url_path='sync-bcv')
    def sync_bcv(self, request):
        """Dispara la sincronización con el BCV."""
        try:
            results = BCVRateService.fetch_and_update_rates()
            return Response(results)
        except Exception as e:
            return Response(
                {'error': str(e)}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
class LatestExchangeRateView(APIView):

    """
    Devuelve la tasa de cambio más reciente para una moneda dada.
    Si no existe tasa para el día de hoy, intenta obtenerla del BCV.
    """
    def get(self, request):
        currency_code = request.query_params.get('currency', 'USD')
        today = timezone.now().date()

        # 1. Intentar buscar en DB la tasa de HOY
        rate_obj = ExchangeRate.objects.filter(
            currency__code=currency_code,
            date_valid__date=today
        ).first()

        # 2. Si no existe la de hoy, llamamos al Servicio BCV (Scraping)
        if not rate_obj:
            print(f"🔄 Tasa {currency_code} desactualizada. Consultando BCV...")
            try:
                # Tu servicio guarda en DB, así que solo lo ejecutamos
                results = BCVRateService.fetch_and_update_rates()
                
                # Verificamos si la operación fue exitosa para la moneda solicitada
                if results.get(currency_code, {}).get('status') == 'success':
                    # Volvemos a consultar la DB recién actualizada
                    rate_obj = ExchangeRate.objects.filter(
                        currency__code=currency_code
                    ).order_by('-date_valid').first()
            except Exception as e:
                print(f"Error consultando servicio BCV: {e}")

        # 3. Si aún no hay tasa (ej. BCV caído y DB vacía), buscar la última histórica
        if not rate_obj:
             rate_obj = ExchangeRate.objects.filter(
                currency__code=currency_code
            ).order_by('-date_valid').first()

        # 4. Respuesta final
        if rate_obj:
            return Response({
                "currency": currency_code,
                "rate": rate_obj.rate,
                "date": rate_obj.date_valid,
                "source": rate_obj.source
            })
        else:
            # Fallback extremo si no hay nada en DB ni internet
            return Response(
                {"currency": currency_code, "rate": 60.00, "note": "Valor por defecto"}, 
                status=status.HTTP_200_OK
            )
class LaborContractViewSet(viewsets.ModelViewSet):
    queryset = LaborContract.objects.all()
    serializer_class = LaborContractSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['employee', 'is_active', 'branch']

class DepartmentViewSet(viewsets.ModelViewSet):
    queryset = Department.objects.all()
    serializer_class = DepartmentSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    search_fields = ['name']
    search_fields = ['name']
    filterset_fields = ['branch']

class JobPositionViewSet(viewsets.ModelViewSet):
    queryset = JobPosition.objects.all()
    serializer_class = JobPositionSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    search_fields = ['name', 'code']
    filterset_fields = ['department']


class EmployeeViewSet(viewsets.ModelViewSet):
    queryset = Employee.objects.all()
    serializer_class = EmployeeSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    search_fields = ['first_name', 'last_name', 'national_id']
    filterset_fields = ['is_active', 'branch']
    @action(detail=True, methods=['get', 'post'], url_path='simulate-payslip')
    def simulate_payslip(self, request, pk=None):
        """
        GET/POST /api/employees/{id}/simulate-payslip/
        Calcula la nómina proyectada usando el Engine Orquestador.
        """
        employee = self.get_object()
        # 1. Obtener variables de entrada (POST json o GET params)
        # 1. Obtener variables de entrada (POST json o GET params)
        input_variables = None
        period_id = None
        
        if request.method == 'POST':
            data = request.data.copy()
            period_id = data.pop('period_id', None)
            if data:
                input_variables = data
        else:
            # Convertir query params a dict
            data = {k: v for k, v in request.query_params.items()}
            period_id = data.pop('period_id', None)
            if data:
                input_variables = data

        # 2. Buscar Contrato Activo
        contract = LaborContract.objects.filter(employee=employee, is_active=True).first()
        if not contract:
            return Response(
                {"error": "El empleado no tiene contrato activo."}, 
                status=status.HTTP_400_BAD_REQUEST
            )
            
        # 2.1 Buscar Periodo (si se indica)
        period = None
        if period_id:
            try:
                period = PayrollPeriod.objects.get(pk=period_id)
            except PayrollPeriod.DoesNotExist:
                pass

        try:
            # 3. Inicializar Motor
            # Si input_variables es None, el motor cargará las novedades de DB (si hay periodo)
            payment_date = period.payment_date if period else timezone.now().date()
            
            engine = PayrollEngine(
                contract=contract, 
                period=period,
                payment_date=payment_date, 
                input_variables=input_variables
            )

            # 4. Delegar el cálculo al Engine (Orquestación)
            # Esto automáticamente inyecta Sueldo Base, IVSS, FAOV y Novedades
            result = engine.calculate_payroll()

            # 5. Forzar conversión de Decimals a Float para facilitar el consumo en JS formatCurrency
            result['totals'] = {k: float(v) if isinstance(v, Decimal) else v for k, v in result.get('totals', {}).items()}
            for line in result.get('lines', []):
                for k, v in line.items():
                    if isinstance(v, Decimal):
                        line[k] = float(v)

            # 5. Retornar el resultado directamente (Ya viene formateado en JSON)
            return Response(result)

        except Exception as e:
            # Log de error para debug en consola
            print(f"Error calculando simulación: {e}")
            return Response(
                {"error": f"Error interno de cálculo: {str(e)}"}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['get'], url_path='payment-history')
    def payment_history(self, request, pk=None):
        """
        GET /api/employees/{id}/payment-history/
        Retorna un historial unificado de pagos (Nómina, Vacaciones, Anticipos, Liquidación).
        """
        employee = self.get_object()
        history = []

        # 1. Nómina Regular (PayrollReceipt)
        receipts = PayrollReceipt.objects.filter(employee=employee).select_related('period').order_by('-period__payment_date')
        for r in receipts:
            history.append({
                'id': f'payroll_{r.id}',
                'original_id': r.id,
                'type': 'PAYROLL',
                'type_label': 'Nómina',
                'date': r.period.payment_date if r.period else r.created_at.date(),
                'description': (r.period.name if r.period else f'Periodo {r.period_id} (Eliminado)'),
                'amount_ves': r.net_pay_ves,
                'amount_usd': getattr(r, 'net_pay_usd_ref', Decimal('0.00')),
                'download_url': f'payslips/{r.id}/export-pdf/'
            })

        # 2. Vacaciones (VacationPayment)
        from django.db.models import Prefetch
        vacations = VacationPayment.objects.filter(vacation_request__employee=employee).select_related('vacation_request').order_by('-payment_date')
        for v in vacations:
            req = v.vacation_request
            history.append({
                'id': f'vacation_{v.id}',
                'original_id': v.id,
                'type': 'VACATION',
                'type_label': 'Vacaciones',
                'date': v.payment_date,
                'description': f'Disfrute {req.start_date} al {req.end_date} ({req.days_requested} días)',
                'amount_ves': v.net_amount_ves,
                'amount_usd': (v.net_amount_ves / v.exchange_rate) if v.exchange_rate and v.exchange_rate > 0 else Decimal('0.00'),
                'download_url': f'vacations/{req.id}/export-pdf/' 
            })

        # 3. Prestaciones - Anticipos (SocialBenefitsLedger)
        advances = SocialBenefitsLedger.objects.filter(
            employee=employee, 
            transaction_type=SocialBenefitsLedger.TransactionType.ANTICIPO
        ).order_by('-transaction_date')
        for a in advances:
            # Los anticipos están como cargos (negativos) en el ledger, los mostramos positivos como pago
            amount = abs(a.amount) if a.amount else Decimal('0.00')
            history.append({
                'id': f'advance_{a.id}',
                'original_id': a.id,
                'type': 'BENEFITS_ADVANCE',
                'type_label': 'Anticipo Prestaciones',
                'date': a.transaction_date,
                'description': a.period_description or 'Anticipo',
                'amount_ves': amount,
                'amount_usd': Decimal('0.00'), # Suponiendo VES para anticipos
                'download_url': None # Sin URL de recibo PDF por ahora
            })

        # 4. Liquidaciones (SocialBenefitsSettlement)
        # Identificando empleado por field employee_national_id ya que no hay ForeignKey directa a Employee, 
        # o a través del contract
        settlements = SocialBenefitsSettlement.objects.filter(
            contract__employee=employee,
            status=SocialBenefitsSettlement.SettlementStatus.PAID
        ).order_by('-settlement_date')
        
        for s in settlements:
            history.append({
                'id': f'settlement_{s.id}',
                'original_id': s.id,
                'type': 'SETTLEMENT',
                'type_label': 'Liquidación Final',
                'date': s.settlement_date,
                'description': 'Cierre de Relación Laboral',
                'amount_ves': s.settlement_amount,
                'amount_usd': Decimal('0.00'),
                'download_url': f'contracts/{s.contract_id}/export-settlement-pdf/'
            })

        # Ordenar todo por fecha descendente
        history.sort(key=lambda x: x['date'], reverse=True)

        return Response(history)

    def update(self, request, *args, **kwargs):
        partial = True # Forzamos parcial para que no borre datos no enviados
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        return Response(serializer.data)
    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        try:
            self.perform_destroy(instance)
            return Response(status=status.HTTP_204_NO_CONTENT)
        except ProtectedError:
            # Capturamos el error de integridad y devolvemos un mensaje amigable
            return Response(
                {"error": "No se puede eliminar este colaborador porque tiene histórico de nóminas o contratos. Proceda a desactivarlo (Dar de baja)."},
                status=status.HTTP_400_BAD_REQUEST
            )

class PayrollPeriodViewSet(viewsets.ModelViewSet):
    queryset = PayrollPeriod.objects.all()
    serializer_class = PayrollPeriodSerializer

    @action(detail=True, methods=['get', 'post'], url_path='preview-payroll')
    def preview_payroll(self, request, pk=None):
        """
        GET/POST /api/payroll-periods/{id}/preview-payroll/
        Calcula la nómina proyectada para todos los empleados del periodo.
        """
        try:
            from ..services.payroll import PayrollProcessor
            manual_rate = request.data.get('manual_rate') if request.method == 'POST' else request.query_params.get('manual_rate')
            result = PayrollProcessor.preview_period(pk, manual_rate=manual_rate)
            return Response(result, status=status.HTTP_200_OK)
        except ValueError as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"error": f"Falla inesperada en previsualización: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['post'], url_path='close-period')
    def close_period(self, request, pk=None):
        """
        POST /api/payroll-periods/{id}/close-period/
        Inicia el proceso de cálculo masivo y cierre inmutable.
        """
        try:
            from ..services import PayrollProcessor
            manual_rate = request.data.get('manual_rate')
            result = PayrollProcessor.process_period(pk, user=request.user, manual_rate=manual_rate)
            return Response(result, status=status.HTTP_200_OK)
        except ValueError as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"error": f"Falla inesperada en cierre: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['get'], url_path='export-pdf')
    def export_pdf(self, request, pk=None):
        """
        Genera el PDF masivo de recibos del periodo.
        """
        period = self.get_object()
        tipo_recibo = request.query_params.get('tipo', 'todos')
        
        # Obtener recibos asociados a este periodo
        payslips_qs = period.receipts.select_related('employee', 'employee__department').prefetch_related('lines').all()
        
        if not payslips_qs.exists():
            return Response({"error": "No hay recibos generados para este periodo."}, status=404)

        # 1. Obtener conceptos configurados para el recibo (Estructura Fija)
        configured_concepts = PayrollConcept.objects.filter(
            appears_on_receipt=True
        ).order_by('receipt_order')
        
        # Mapeo de códigos configurados para referencia rápida
        configured_codes = {c.code: c for c in configured_concepts}

        # 2. Pre-procesar cada payslip
        processed_payslips = []
        for payslip in payslips_qs:
            actual_details = {d.concept_code: d for d in payslip.lines.all()}
            used_codes = set()
            payslip_rows = []
            
            # Primero, agregar conceptos configurados (Orden Fijo)
            for concept in configured_concepts:
                detail = actual_details.get(concept.code)
                used_codes.add(concept.code)
                
                if detail:
                    row = {
                        'name': detail.concept_name,
                        'concept_name': detail.concept_name, # Alias para compatibilidad
                        'code': detail.concept_code,
                        'kind': detail.kind,
                        'quantity': getattr(detail, 'quantity', 0) or 0,
                        'unit': getattr(detail, 'unit', '') or '',
                        'amount_ves': detail.amount_ves,
                        'tipo_recibo': detail.tipo_recibo,
                    }
                elif concept.show_even_if_zero:
                    row = {
                        'name': concept.name,
                        'concept_name': concept.name, # Alias para compatibilidad
                        'code': concept.code,
                        'kind': concept.kind,
                        'quantity': 0,
                        'unit': '',
                        'amount_ves': Decimal('0.00'),
                        'tipo_recibo': 'salario',
                    }
                else:
                    # Si no hay detalle y no se fuerza el cero, no agregar
                    used_codes.remove(concept.code) # Marcar como no usado realmente
                    continue
                
                row['earning_amount'] = row['amount_ves'] if row['kind'] == 'EARNING' else Decimal('0.00')
                row['deduction_amount'] = row['amount_ves'] if row['kind'] == 'DEDUCTION' else Decimal('0.00')
                payslip_rows.append(row)
            
            # Segundo, agregar detalles calculados que NO estaban en la configuración fija (Fallback)
            for code, detail in actual_details.items():
                if code not in used_codes:
                    row = {
                        'name': detail.concept_name,
                        'concept_name': detail.concept_name,
                        'code': detail.concept_code,
                        'kind': detail.kind,
                        'quantity': getattr(detail, 'quantity', 0) or 0,
                        'unit': getattr(detail, 'unit', '') or '',
                        'amount_ves': detail.amount_ves,
                        'tipo_recibo': detail.tipo_recibo,
                        'earning_amount': detail.amount_ves if detail.kind == 'EARNING' else Decimal('0.00'),
                        'deduction_amount': detail.amount_ves if detail.kind == 'DEDUCTION' else Decimal('0.00'),
                    }
                    payslip_rows.append(row)
            
            payslip.processed_rows = payslip_rows
            processed_payslips.append(payslip)

        # Contexto para el Template HTML
        company_config = Company.objects.first()
        context = {
            'payslips': processed_payslips,
            'period_name': period.name,
            'start_date': period.start_date,
            'end_date': period.end_date,
            'payment_date': period.payment_date,
            'tenant_name': request.tenant.name if hasattr(request, 'tenant') else "Nóminix",
            'tenant_rif': request.tenant.rif if hasattr(request, 'tenant') else "",
            'tenant_address': request.tenant.address if hasattr(request, 'tenant') else "",
            'company_config': company_config,
            'tipo_recibo': tipo_recibo,
        }

        # Seleccionar template según tipo de recibo
        template_map = {
            'todos': 'payroll/payslip_batch.html',
            'salario': 'payroll/recibo_salario.html',
            'complemento': 'payroll/recibo_complemento.html',
            'cestaticket': 'payroll/recibo_cestaticket.html',
        }
        template_name = template_map.get(tipo_recibo, 'payroll/payslip_batch.html')

        # 1. Renderizar HTML string
        html_string = render_to_string(template_name, context)

        # 2. Generar PDF con WeasyPrint
        pdf_file = weasyprint.HTML(string=html_string).write_pdf()

        # 3. Retornar respuesta HTTP con Content-Type correcto
        suffix = f'_{tipo_recibo}' if tipo_recibo != 'todos' else ''
        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="nomina{suffix}_{period.id}.pdf"'
        return response

    @action(detail=True, methods=['get'], url_path='export-finance')
    def export_finance_report(self, request, pk=None):
        """
        Genera un reporte Excel (.xlsx) nativo para Finanzas.
        Permite definir tipos de celda explícitos.
        """
        period = self.get_object()
        # Usamos select_related para optimizar la DB
        payslips = period.receipts.select_related('employee').all()

        if not payslips.exists():
            return Response({"error": "No hay recibos generados para este periodo."}, status=404)

        # 1. Configuración del archivo Excel
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="finanzas_nomina_{period.id}.xlsx"'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte Finanzas"

        # 2. Encabezados
        headers = ['Cédula', 'Empleado', 'Sede', 'Banco', 'Tipo Cuenta', 'Número de Cuenta', 'Monto VES', 'Tasa BCV', 'Monto USD Ref']
        ws.append(headers)

        # Estilo para encabezados (Negrita)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # 3. Iterar datos
        for p in payslips:
            emp = p.employee
            
            # Cálculos previos seguros
            monto_ves = p.net_pay_ves or Decimal(0)
            tasa = p.exchange_rate_snapshot or Decimal(0)
            monto_usd = (monto_ves / tasa).quantize(Decimal('0.01')) if tasa > 0 else Decimal(0)

            # Preparamos la fila
            row = [
                emp.national_id,
                emp.full_name,
                emp.branch.name if emp.branch else 'N/D',
                emp.bank_name or 'N/D',
                emp.bank_account_type or 'N/D',
                emp.bank_account_number or 'N/D', # Se inserta como string inicialmente
                monto_ves,
                tasa,
                monto_usd
            ]
            ws.append(row)
            
            # 4. DEFINICIÓN DE FORMATOS POR CELDA (La magia ocurre aquí)
            current_row = ws.max_row
            
            # Columna F (Número de Cuenta): Forzar formato Texto para evitar notación científica
            cell_acc = ws.cell(row=current_row, column=6)
            cell_acc.number_format = '@'  # '@' significa Texto en Excel
            
            # Columna G (Monto VES): Formato numérico con separadores
            cell_ves = ws.cell(row=current_row, column=7)
            cell_ves.number_format = '#,##0.00' 

            # Columna H (Tasa): Formato con más decimales si es necesario
            cell_rate = ws.cell(row=current_row, column=8)
            cell_rate.number_format = '#,##0.0000'

            # Columna I (USD): Formato numérico
            cell_usd = ws.cell(row=current_row, column=9)
            cell_usd.number_format = '#,##0.00'

        # 5. Ajustar ancho de columnas automáticamente (Opcional pero útil)
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # 6. Guardar en la respuesta
        wb.save(response)
        return response

class PayrollReceiptViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet solo lectura para consultar recibos históricos.
    """
    queryset = PayrollReceipt.objects.all()
    serializer_class = PayrollReceiptSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['period', 'employee', 'currency_code']

    @action(detail=True, methods=['get'], url_path='export-pdf')
    def export_pdf(self, request, pk=None):
        """
        Genera el PDF individual para un recibo de pago.
        """
        import weasyprint
        from django.template.loader import render_to_string
        from django.http import HttpResponse

        payslip = self.get_object()
        
        # 1. Obtener conceptos configurados
        configured_concepts = PayrollConcept.objects.filter(
            appears_on_receipt=True
        ).order_by('receipt_order')

        # 2. Pre-procesar filas
        actual_details = {d.concept_code: d for d in payslip.lines.all()}
        used_codes = set()
        payslip_rows = []
        
        # Primero, conceptos configurados
        for concept in configured_concepts:
            detail = actual_details.get(concept.code)
            used_codes.add(concept.code)
            
            if detail:
                row = {
                    'name': detail.concept_name,
                    'code': detail.concept_code,
                    'kind': detail.kind,
                    'quantity': getattr(detail, 'quantity', 0) or 0,
                    'unit': getattr(detail, 'unit', '') or '',
                    'amount_ves': detail.amount_ves,
                    'tipo_recibo': detail.tipo_recibo,
                }
            elif concept.show_even_if_zero:
                row = {
                    'name': concept.name,
                    'code': concept.code,
                    'kind': concept.kind,
                    'quantity': 0,
                    'unit': '',
                    'amount_ves': Decimal('0.00'),
                    'tipo_recibo': 'salario',
                }
            else:
                used_codes.remove(concept.code)
                continue
            
            row['earning_amount'] = row['amount_ves'] if row['kind'] == 'EARNING' else None
            row['deduction_amount'] = row['amount_ves'] if row['kind'] == 'DEDUCTION' else None
            payslip_rows.append(row)

        # Segundo, fallback para detalles no configurados
        for code, detail in actual_details.items():
            if code not in used_codes:
                row = {
                    'name': detail.concept_name,
                    'code': detail.concept_code,
                    'kind': detail.kind,
                    'quantity': getattr(detail, 'quantity', 0) or 0,
                    'unit': getattr(detail, 'unit', '') or '',
                    'amount_ves': detail.amount_ves,
                    'tipo_recibo': detail.tipo_recibo,
                    'earning_amount': detail.amount_ves if detail.kind == 'EARNING' else None,
                    'deduction_amount': detail.amount_ves if detail.kind == 'DEDUCTION' else None,
                }
                payslip_rows.append(row)
        
        payslip.processed_rows = payslip_rows
        company_config = Company.objects.first()
        
        context = {
            'payslips': [payslip],
            'period_name': payslip.period.name,
            'start_date': payslip.period.start_date,
            'end_date': payslip.period.end_date,
            'payment_date': payslip.period.payment_date,
            'tenant_name': request.tenant.name if hasattr(request, 'tenant') else "Nóminix",
            'tenant_rif': request.tenant.rif if hasattr(request, 'tenant') else "",
            'tenant_address': request.tenant.address if hasattr(request, 'tenant') else "",
            'company_config': company_config,
        }

        html_string = render_to_string('payroll/payslip_batch.html', context)
        pdf_file = weasyprint.HTML(string=html_string).write_pdf()

        response = HttpResponse(pdf_file, content_type='application/pdf')
        filename = f"recibo_{payslip.employee.national_id}_{payslip.period.id}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response



class PayrollNoveltyViewSet(viewsets.ModelViewSet):
    """
    Gestión de incidencias de nómina (Novedades).
    """
    queryset = PayrollNovelty.objects.all()
    serializer_class = PayrollNoveltySerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['employee', 'period', 'concept_code']

    @action(detail=False, methods=['get'], url_path='metadata')
    def metadata(self, request):
        """
        GET /api/payroll-novelties/metadata/
        Devuelve la lista de conceptos que pueden cargarse como novedades.
        """
        # Conceptos que son fórmulas dinámicas o préstamos suelen ser novedades
        # También cualquier cosa que el usuario quiera marcar específicamente (is_system=False)
        concepts = PayrollConcept.objects.filter(
            active=True
        ).exclude(
            behavior__in=['SALARY_BASE', 'CESTATICKET', 'COMPLEMENT', 'LAW_DEDUCTION']
        ).order_by('receipt_order')

        # Si el usuario quiere ver los de sistema que son novedades (Faltas, etc)
        # los incluimos explícitamente si cumplen con el criterio anterior.
        
        data = []
        for c in concepts:
            data.append({
                'code': c.code,
                'name': c.name,
                'kind': c.kind,
                'is_system': c.is_system,
                'behavior': c.behavior
            })
            
        return Response({
            'concepts': data,
            'mappings': PayrollEngine.NOVELTY_MAP if hasattr(PayrollEngine, 'NOVELTY_MAP') else {}
        })

    @action(detail=False, methods=['post'], url_path='batch')
    def batch(self, request):
        """
        POST /api/payroll-novelties/batch/
        Carga masiva de novedades.
        Recibe: [{"employee_id": 1, "period_id": 5, "concept_code": "H_EXTRA", "amount": 4}, ...]
        """
        data = request.data
        if not isinstance(data, list):
            return Response(
                {"error": "Se esperaba una lista de novedades (array)."}, 
                status=status.HTTP_400_BAD_REQUEST
            )
            
        processed_ids = []
        try:
            with transaction.atomic():
                for item in data:
                    # Usamos update_or_create para manejar ediciones rápidas en el grid
                    obj, created = PayrollNovelty.objects.update_or_create(
                        employee_id=item.get('employee_id'),
                        period_id=item.get('period_id'),
                        concept_code=item.get('concept_code'),
                        defaults={'amount': Decimal(str(item.get('amount', 0)))}
                    )
                    processed_ids.append(obj.id)
            
            return Response({
                "status": "success", 
                "processed_count": len(processed_ids)
            }, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            return Response(
                {"error": f"Falla en procesamiento batch: {str(e)}"}, 
                status=status.HTTP_400_BAD_REQUEST
            )
class CompanyConfigView(views.APIView):
    """
    Endpoint único para obtener/editar la configuración de la empresa.
    Siempre devuelve el primer objeto encontrado o crea uno por defecto.
    """
    def get(self, request):
        company = Company.objects.first()
        if not company:
            # Crear datos default si no existen
            company = Company.objects.create(name="Mi Empresa C.A.", rif="J-00000000-0")
        serializer = CompanySerializer(company)
        return response.Response(serializer.data)

    def put(self, request):
        company = Company.objects.first()
        serializer = CompanySerializer(company, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return response.Response(serializer.data)
        return response.Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class LoanViewSet(viewsets.ModelViewSet):
    queryset = Loan.objects.all()
    serializer_class = LoanSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['employee', 'status', 'currency', 'frequency']
    search_fields = ['employee__first_name', 'employee__last_name', 'employee__national_id', 'description']

class LoanPaymentViewSet(viewsets.ModelViewSet):
    queryset = LoanPayment.objects.all()
    serializer_class = LoanPaymentSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['loan', 'payslip']

class PayrollVariablesView(APIView):
    """
    Devuelve el inventario de variables disponibles para fórmulas.
    """
    def get(self, request):
        return Response(PayrollEngine.get_variable_inventory())

class ValidateFormulaView(APIView):
    """
    Valida una fórmula y devuelve el resultado y la traza.
    """
    def post(self, request):
        formula = request.data.get('formula')
        context = request.data.get('context', {})
        if not formula:
            return Response({"error": "Fórmula no proporcionada"}, status=status.HTTP_400_BAD_REQUEST)
        
        result = PayrollEngine.validate_formula(formula, context)
        return Response(result)


# =============================================================================
# SOCIAL BENEFITS VIEWSET (Prestaciones Sociales)
# =============================================================================

class SocialBenefitsViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet para gestión de Prestaciones Sociales.
    
    Endpoints:
    - GET /api/social-benefits/ - Lista movimientos del ledger
    - GET /api/social-benefits/{id}/ - Detalle de un movimiento
    - POST /api/social-benefits/run-quarterly/ - Procesar garantía trimestral
    - GET /api/social-benefits/simulate-settlement/?contract_id=X - Simular liquidación
    - POST /api/social-benefits/request-advance/ - Solicitar anticipo
    """
    queryset = SocialBenefitsLedger.objects.all()
    serializer_class = SocialBenefitsLedgerSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['employee', 'contract', 'transaction_type']

    @action(detail=False, methods=['post'], url_path='run-quarterly')
    def run_quarterly(self, request):
        """
        POST /api/social-benefits/run-quarterly/
        
        Procesa el abono de garantía trimestral (15 días de salario integral).
        
        Request Body:
        {
            "contract_id": 1,
            "period_description": "Q1-2026",
            "transaction_date": "2026-03-31" (opcional)
        }
        """
        serializer = QuarterlyGuaranteeSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        data = serializer.validated_data
        contract_id = data['contract_id']
        period_description = data['period_description']
        transaction_date = data.get('transaction_date') or timezone.now().date()
        
        try:
            contract = LaborContract.objects.get(pk=contract_id)
        except LaborContract.DoesNotExist:
            return Response(
                {'error': f'Contrato con ID {contract_id} no encontrado'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Importar el motor de prestaciones sociales
        from ..services.social_benefits_engine import process_quarterly_guarantee
        
        try:
            # Obtener información del usuario para auditoría
            created_by = getattr(request.user, 'username', 'API')
            ip_address = self._get_client_ip(request)
            
            entry = process_quarterly_guarantee(
                contract=contract,
                transaction_date=transaction_date,
                period_description=period_description,
                created_by=created_by,
                ip_address=ip_address,
            )
            
            return Response(
                SocialBenefitsLedgerSerializer(entry).data,
                status=status.HTTP_201_CREATED
            )
        except Exception as e:
            return Response(
                {'error': f'Error procesando garantía trimestral: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'], url_path='simulate-settlement')
    def simulate_settlement(self, request):
        """
        GET /api/social-benefits/simulate-settlement/?contract_id=X
        
        Simula la liquidación final comparando Garantía vs Retroactivo.
        NO persiste datos, solo retorna el cálculo.
        """
        contract_id = request.query_params.get('contract_id')
        termination_date_str = request.query_params.get('termination_date')
        
        if not contract_id:
            return Response(
                {'error': 'Se requiere contract_id como parámetro'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            contract = LaborContract.objects.get(pk=contract_id)
        except LaborContract.DoesNotExist:
            return Response(
                {'error': f'Contrato con ID {contract_id} no encontrado'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Parsear fecha de terminación o usar hoy
        if termination_date_str:
            from datetime import datetime
            try:
                termination_date = datetime.strptime(termination_date_str, '%Y-%m-%d').date()
            except ValueError:
                return Response(
                    {'error': 'Formato de fecha inválido. Use YYYY-MM-DD'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        else:
            termination_date = timezone.now().date()
        
        # Importar y ejecutar el cálculo
        from ..services.social_benefits_engine import calculate_final_settlement
        
        try:
            comparison = calculate_final_settlement(contract, termination_date)
            
            # Convertir Decimals a float para JSON
            result = {k: float(v) if isinstance(v, Decimal) else v for k, v in comparison.items()}
            
            return Response(result)
        except Exception as e:
            return Response(
                {'error': f'Error calculando liquidación: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'], url_path='export-simulation-pdf')
    def export_simulation_pdf(self, request):
        """
        GET /api/social-benefits/export-simulation-pdf/?contract_id=X
        Genera el cuadro comparativo en formato PDF.
        """
        import weasyprint
        from django.template.loader import render_to_string
        from django.http import HttpResponse

        contract_id = request.query_params.get('contract_id')
        termination_date_str = request.query_params.get('termination_date')
        
        if not contract_id:
            return Response({'error': 'Se requiere contract_id'}, status=status.HTTP_400_BAD_REQUEST)
            
        try:
            contract = LaborContract.objects.get(pk=contract_id)
        except LaborContract.DoesNotExist:
            return Response({'error': 'Contrato no encontrado'}, status=status.HTTP_404_NOT_FOUND)
            
        if termination_date_str:
            from datetime import datetime
            try:
                termination_date = datetime.strptime(termination_date_str, '%Y-%m-%d').date()
            except ValueError:
                return Response({'error': 'Formato de fecha inválido. YYYY-MM-DD'}, status=status.HTTP_400_BAD_REQUEST)
        else:
            termination_date = timezone.now().date()
            
        from ..services.social_benefits_engine import calculate_final_settlement
        
        try:
            comparison = calculate_final_settlement(contract, termination_date)
            
            context = {
                'employee': contract.employee,
                'contract': contract,
                'comparison': comparison,
                'termination_date': termination_date,
                'simulation_date': timezone.now().date(),
                'tenant_name': request.tenant.name if hasattr(request, 'tenant') else "Empresa Demo",
                'tenant_rif': request.tenant.rif if hasattr(request, 'tenant') else "J-00000000-0",
            }

            html_string = render_to_string('payroll/cuadro_liquidacion.html', context)
            pdf_file = weasyprint.HTML(string=html_string).write_pdf()

            response = HttpResponse(pdf_file, content_type='application/pdf')
            filename = f"cuadro_liquidacion_{contract.employee.national_id}.pdf"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['post'], url_path='request-advance')
    def request_advance(self, request):
        """
        POST /api/social-benefits/request-advance/
        
        Solicita un anticipo de prestaciones sociales.
        
        Restricción de Negocio: El monto no puede exceder el 75% del saldo actual.
        
        Request Body:
        {
            "contract_id": 1,
            "amount": 1000.00,
            "notes": "Anticipo personal" (opcional)
        }
        """
        serializer = AdvanceRequestSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        data = serializer.validated_data
        contract_id = data['contract_id']
        amount = data['amount']
        notes = data.get('notes', '')
        
        try:
            contract = LaborContract.objects.get(pk=contract_id)
        except LaborContract.DoesNotExist:
            return Response(
                {'error': f'Contrato con ID {contract_id} no encontrado'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        employee = contract.employee
        
        # Importar helper para obtener saldo actual
        from ..services.social_benefits_engine import (
            get_current_balance,
            calculate_comprehensive_salary
        )
        
        # 1. Obtener saldo actual
        current_balance = get_current_balance(employee)
        
        # 2. Validar restricción de negocio: máximo 75% del saldo
        max_allowed = current_balance * Decimal('0.75')
        
        if amount > max_allowed:
            return Response(
                {
                    'error': f'El monto solicitado ({amount}) excede el 75% del saldo disponible.',
                    'current_balance': float(current_balance),
                    'max_allowed': float(max_allowed.quantize(Decimal('0.01'))),
                },
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if amount <= Decimal('0'):
            return Response(
                {'error': 'El monto debe ser mayor a cero'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # 3. Calcular salario integral para el snapshot
        salary_result = calculate_comprehensive_salary(contract, timezone.now().date())
        daily_salary_used = salary_result['daily_salary_integral']
        
        # 4. Crear el registro de ANTICIPO (monto negativo)
        created_by = getattr(request.user, 'username', 'API')
        ip_address = self._get_client_ip(request)
        
        try:
            anticipo_entry = SocialBenefitsLedger(
                employee=employee,
                contract=contract,
                transaction_type=SocialBenefitsLedger.TransactionType.ANTICIPO,
                transaction_date=timezone.now().date(),
                period_description=f'Anticipo {timezone.now().strftime("%b %Y")}',
                # Snapshot de auditoría
                basis_days=Decimal('0'),  # No aplica para anticipos
                daily_salary_used=daily_salary_used,
                previous_balance=current_balance,
                # Monto NEGATIVO (es un cargo, no un abono)
                amount=-amount,
                balance=current_balance - amount,
                # Trazabilidad
                calculation_formula='Anticipo de Prestaciones (-amount)',
                calculation_trace=f'Saldo anterior: {current_balance}, Anticipo: -{amount}, Nuevo saldo: {current_balance - amount}',
                # Auditoría
                created_by=created_by,
                ip_address=ip_address,
                notes=notes,
            )
            anticipo_entry.save()
            
            return Response(
                SocialBenefitsLedgerSerializer(anticipo_entry).data,
                status=status.HTTP_201_CREATED
            )
        except Exception as e:
            return Response(
                {'error': f'Error creando anticipo: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'], url_path='balance')
    def get_balance(self, request):
        """
        GET /api/social-benefits/balance/?employee_id=X
        
        Obtiene el saldo actual de prestaciones de un empleado.
        """
        employee_id = request.query_params.get('employee_id')
        
        if not employee_id:
            return Response(
                {'error': 'Se requiere employee_id como parámetro'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            employee = Employee.objects.get(pk=employee_id)
        except Employee.DoesNotExist:
            return Response(
                {'error': f'Empleado con ID {employee_id} no encontrado'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        from ..services.social_benefits_engine import get_current_balance
        
        balance = get_current_balance(employee)
        
        return Response({
            'employee_id': employee.id,
            'employee_name': employee.full_name,
            'balance': float(balance),
        })

    def _get_client_ip(self, request):
        """Extrae la IP del cliente del request."""
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            return x_forwarded_for.split(',')[0].strip()
        return request.META.get('REMOTE_ADDR')


class InterestRateBCVViewSet(viewsets.ModelViewSet):
    """
    CRUD para las tasas de interés del BCV.
    """
    queryset = InterestRateBCV.objects.all()
    serializer_class = InterestRateBCVSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['year', 'month']

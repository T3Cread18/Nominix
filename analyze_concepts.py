import openpyxl
import os

file_path = 'Copia de NOMINA PARA PABLO.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=False)
ws = wb['CALCULO BS']

with open('concepts_output.txt', 'w', encoding='utf-8') as f:
    f.write(f"--- ANALYZING CONCEPTS: {file_path} ---\n")

    target_headers = [
        'DIAS LABORADOS', 'DIAS DE DESCANSO', 'FERIADO', 
        'REPOSO 100%', 'REPOSO 33,33%', 'TRANSPORTE', 
        'HORAS EXTRAS', 'BONO NOCTURNO', 'BONIFICACION'
    ]

    header_map = {}
    for r in range(5, 9):
        for c in range(1, 60):
            val = ws.cell(row=r, column=c).value
            if val:
                s_val = str(val).strip().upper()
                if any(t in s_val for t in target_headers):
                    header_map[c] = s_val

    f.write(f"Found {len(header_map)} relevant columns.\n\n")

    for c in sorted(header_map.keys()):
        header = header_map[c]
        formula = ws.cell(row=14, column=c).value
        # Check column to the right (often amount if this is just quantity/days)
        val_right = ws.cell(row=14, column=c+1).value
        val_right_2 = ws.cell(row=14, column=c+2).value
        
        f.write(f"COLUMN {c} ({openpyxl.utils.get_column_letter(c)})\n")
        f.write(f"  HEADER: {header}\n")
        f.write(f"  FORMULA (Row 14): {formula}\n")
        f.write(f"  VAL RIGHT (+1): {val_right}\n")
        f.write(f"  VAL RIGHT (+2): {val_right_2}\n")
        f.write("-" * 20 + "\n")

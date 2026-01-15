import os
import datetime
from openpyxl import load_workbook
import json

file_path = 'c:\\Desarrollo\\RRHH\\trabajadores.xlsx'

def clean_text(text, type='other'):
    if not text: return ""
    import re
    text = str(text).upper().strip()
    # Normalize common abbreviations
    text = text.replace("C.A.", "CA").replace("C.A", "CA")
    text = re.sub(r'\s+', ' ', text)
    
    if type == 'branch':
        # Clean branch specific suffixes but keep core identity
        text = text.replace(" /CORPORATIVO", "").replace(" SUCURSAL", "").strip()
    
    if type == 'position':
        # Unify variations of Auxiliar
        if any(x in text for x in ["AUX ", "AUX.", "AUXILIAR"]):
            if "ALMACEN" in text: return "AUXILIAR DE ALMACEN"
            if "ADMINISTRATIVO" in text: return "ASISTENTE ADMINISTRATIVO"
        if "CHOFER" in text: return "CHOFER"
        if "CAJERO" in text: return "CAJERO (A)"
        
    return text.strip()

# Custom Mapping Overrides
MANUAL_MAPPING = {
    "branches": {
        "FARMACIA FARMANOSTRA GUANARE, C.A. /CORPORATIVO": "FARMACIA FARMANOSTRA GUANARE, CA",
        "FARMACIA MAMANICO OSPINO C.A. SUCURSAL": "FARMACIA MAMANICO OSPINO CA"
    }
}

try:
    wb = load_workbook(file_path, data_only=True)
    
    mapping = {
        "branches": {},
        "departments": {},
        "positions": {}
    }

    unique_raw = {
        "branches": set(),
        "departments": set(),
        "positions": set()
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 4: continue
        
        # Data starts at row 4 (index 3)
        for row in rows[3:]:
            # Column 1: Farmacia, 9: Cargo, 10: Departamento
            b = row[1]
            p = row[9]
            d = row[10]
            
            if b: unique_raw["branches"].add(str(b).strip())
            if p: unique_raw["positions"].add(str(p).strip())
            if d: unique_raw["departments"].add(str(d).strip())

    # Generate normalization map
    def generate_norm_map(raw_set, type_key):
        norm_map = {}
        for raw in sorted(list(raw_set)):
            # Check manual override first
            if type_key in MANUAL_MAPPING and raw in MANUAL_MAPPING[type_key]:
                norm_map[raw] = MANUAL_MAPPING[type_key][raw]
            else:
                norm_map[raw] = clean_text(raw, type_key[:-1]) # remove 's'
        return norm_map

    mapping["branches"] = generate_norm_map(unique_raw["branches"], "branches")
    mapping["positions"] = generate_norm_map(unique_raw["positions"], "positions")
    mapping["departments"] = generate_norm_map(unique_raw["departments"], "departments")

    # Look for collisions (multiple RAW pointing to same CLEAN)
    def find_collisions(m):
        rev = {}
        for raw, clean in m.items():
            if clean not in rev: rev[clean] = []
            rev[clean].append(raw)
        return {k: v for k, v in rev.items() if len(v) > 1}

    collisions = {
        "branches": find_collisions(mapping["branches"]),
        "positions": find_collisions(mapping["positions"]),
        "departments": find_collisions(mapping["departments"])
    }

    print("--- STANDARDIZATION REPORT ---")
    
    print("\n[BRANCHES]")
    for clean, raws in collisions["branches"].items():
        print(f"  * Canonical: {clean}")
        for r in raws: print(f"    <- {r}")

    print("\n[POSITIONS]")
    for clean, raws in collisions["positions"].items():
        print(f"  * Canonical: {clean}")
        for r in raws: print(f"    <- {r}")

    print("\n[DEPARTMENTS]")
    for clean, raws in collisions["departments"].items():
        print(f"  * Canonical: {clean}")
        for r in raws: print(f"    <- {r}")

    with open('c:\\Desarrollo\\RRHH\\normalization_map.json', 'w') as f:
        json.dump(mapping, f, indent=2)
    print("\nNormalization map saved to normalization_map.json")

except Exception as e:
    import traceback
    traceback.print_exc()

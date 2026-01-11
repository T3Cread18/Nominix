import openpyxl
import os

file_path = 'Copia de NOMINA PARA PABLO.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=False)
ws = wb['CALCULO BS']

# Based on previous output, headers seem to be around Row 5
header_row = 5

print(f"--- HEADERS in 'CALCULO BS' (Row {header_row}) ---")
for c in range(10, 25):
    val = ws.cell(row=header_row, column=c).value
    # Also check row 4 and 6 just in case of merged cells
    val_above = ws.cell(row=header_row-1, column=c).value
    val_below = ws.cell(row=header_row+1, column=c).value
    
    print(f"Col {c} ({openpyxl.utils.get_column_letter(c)}):")
    print(f"  Row {header_row-1}: {val_above}")
    print(f"  Row {header_row}: {val}")
    print(f"  Row {header_row+1}: {val_below}")

# Check formula at Row 14, Col 16 again
print("\n--- FORMULA CHECK [14, 16] ---")
print(ws.cell(row=14, column=16).value)

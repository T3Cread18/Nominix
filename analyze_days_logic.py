import openpyxl
import os

file_path = 'Copia de NOMINA PARA PABLO.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=False)
ws = wb['CALCULO BS']

print(f"--- ANALYZING DAY LOGIC IN: {file_path} ---")

# 1. Analyze the Counter Row (Row 4 based on previous finding)
row_4 = []
for c in range(1, 35): # Check first 34 columns (days of month?)
    val = ws.cell(row=4, column=c).value
    row_4.append(str(val))

print(f"\n>>> ROW 4 (Day Counter?):")
print(row_4)

# 2. Check for "Monday" counting logic
# IVSS often uses Mondays. Let's scan for formulas using MOD, WEEKDAY, DIASEM
print(f"\n>>> SCANNING FOR WEEKDAY/MONDAY LOGIC:")
found_weekday = False
for r in range(1, 20):
    for c in range(1, 40):
        val = ws.cell(row=r, column=c).value
        if isinstance(val, str) and ('DIASEM' in val or 'WEEKDAY' in val):
            print(f"  FOUND WEEKDAY FORMULA AT [{r},{c}]: {val}")
            found_weekday = True

if not found_weekday:
    print("  No explicit WEEKDAY/DIASEM formulas found in first 20 rows.")

# 3. Analyze Column Headers for "DIAS" or "LUNES"
print(f"\n>>> HEADERS RELATED TO DAYS (Rows 1-8):")
for r in range(1, 9):
    for c in range(1, 50):
        val = ws.cell(row=r, column=c).value
        if val and isinstance(val, str) and any(x in val.upper() for x in ['DIA', 'LUNES', 'FALTAS', 'ASISTENCIA']):
             print(f"  [{r},{c}] {val}")

# 4. Check formula for IVSS deduction (usually involves Mondays)
# IVSS is usually near the end of the calculation block.
# Let's search for "IVSS" header and print the formula below it.
print(f"\n>>> IVSS FORMULA CHECK:")
ivss_col = -1
ivss_row = -1

# Scan for IVSS header
for r in range(1, 10):
    for c in range(1, 50):
        val = ws.cell(row=r, column=c).value
        if val and "IVSS" in str(val).upper():
            ivss_col = c
            ivss_row = r
            print(f"  Found IVSS Header at [{r},{c}]: {val}")
            break
    if ivss_col != -1: break

if ivss_col != -1:
    # Print formula below header
    val_below = ws.cell(row=ivss_row+1, column=ivss_col).value
    print(f"  Formula below IVSS: {val_below}")
    # Try a few rows down where data likely is
    val_data = ws.cell(row=14, column=ivss_col).value # Assuming data starts row 14
    print(f"  Formula at Row 14 (Data): {val_data}")

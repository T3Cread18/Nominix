import openpyxl
import os

file_path = 'Copia de NOMINA PARA PABLO.xlsx'

if not os.path.exists(file_path):
    print("File not found.")
    exit(1)

print(f"--- DEEP ANALYSIS: {file_path} ---")
wb = openpyxl.load_workbook(file_path, data_only=False)

if 'CALCULO BS' not in wb.sheetnames:
    print("Sheet 'CALCULO BS' not found.")
    exit(1)

ws = wb['CALCULO BS']

# Keywords to look for in headers
keywords = ['HORA', 'DIA', 'NOCT', 'EXTRA', 'FERIADO', 'LUNES', 'DOMINGO']
numeric_patterns = ['0.33', '0,33', '1.33', '1,33', '1.5', '1,5', '30%', '0.3']

print(">>> SCANNING HEADERS (Rows 1-15)...")
header_map = {}
for r in range(1, 15):
    row_values = []
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=r, column=c).value
        # Check for numeric factors often put in headers or near them
        if val:
            s_val = str(val).upper()
            row_values.append(f"[{r},{c}]{s_val}")
            
            # Save potential headers
            if any(k in s_val for k in keywords):
                header_map[c] = s_val
                print(f"  FOUND KEYWORD AT [{r},{c}]: {val}")
            
            # Check for literal factors
            if any(p in s_val for p in numeric_patterns):
                 print(f"  FOUND FACTOR AT [{r},{c}]: {val}")
    
    # Print distinct row content if interesting
    # print(row_values)

print("\n>>> SCANNING FORMULAS (Row + 1 of Header likely)...")
# Assuming data starts around row 10 based on previous analysis
data_start_row = 10 

for c, header in header_map.items():
    cell = ws.cell(row=data_start_row, column=c)
    val = cell.value
    print(f"  Under Header '{header}' (Col {c}):")
    print(f"    Value/Formula: {val}")

print("\n>>> SEARCHING FOR '0.33' IN ALL FORMULAS (First 20 columns)")
for r in range(5, 15):
    for c in range(1, 50):
        cell = ws.cell(row=r, column=c)
        val = str(cell.value)
        if any(x in val for x in ['0.33', '0,33', '/3']):
             print(f"  MATCH AT [{r},{c}]: {val}")


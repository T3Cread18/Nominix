import openpyxl
import os

file_path = 'Copia de NOMINA PARA PABLO.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=False)
ws = wb['CALCULO BS']

c = 35 # Column AI
print(f"--- HEADER CHECK COL {c} (AI) ---")
for r in range(4, 10):
    val = ws.cell(row=r, column=c).value
    print(f"Row {r}: {val}")

print(f"\n--- DATA CHECK COL {c} (AI) ---")
print(f"Row 14: {ws.cell(row=14, column=c).value}")

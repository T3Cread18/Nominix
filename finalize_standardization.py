import os
import json
import datetime
from openpyxl import load_workbook

# Paths
INPUT_EXCEL = 'c:\\Desarrollo\\RRHH\\trabajadores.xlsx'
MAP_FILE = 'c:\\Desarrollo\\RRHH\\normalization_map.json'
OUTPUT_JSON = 'c:\\Desarrollo\\RRHH\\standardized_workers.json'

def clean_cid(val):
    if not val: return None
    # Remove dots and non-numeric chars for the core number
    import re
    cleaned = re.sub(r'[^0-9]', '', str(val))
    return cleaned

try:
    with open(MAP_FILE, 'r') as f:
        mapping = json.load(f)

    wb = load_workbook(INPUT_EXCEL, data_only=True)
    
    final_data = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 4: continue
        
        for i, row in enumerate(rows[3:]):
            # Map raw values to standardized ones
            raw_branch = str(row[1]).strip() if row[1] else None
            raw_pos = str(row[9]).strip() if row[9] else None
            raw_dept = str(row[10]).strip() if row[10] else None
            
            # Skip rows where name is totally empty
            if not row[4] and not row[6]: continue

            worker = {
                "source_sheet": sheet_name,
                "source_row": i + 4,
                "branch": mapping["branches"].get(raw_branch, raw_branch),
                "name": str(row[4]).strip().upper() if row[4] else "SIN NOMBRE",
                "cid": clean_cid(row[6]),
                "raw_cid": str(row[6]),
                "birth_date": row[5].isoformat() if isinstance(row[5], (datetime.date, datetime.datetime)) else str(row[5]),
                "hire_date": row[8].isoformat() if isinstance(row[8], (datetime.date, datetime.datetime)) else str(row[8]),
                "position": mapping["positions"].get(raw_pos, raw_pos),
                "department": mapping["departments"].get(raw_dept, raw_dept),
                "phone": str(row[11]).strip() if row[11] else None,
                "email": str(row[12]).strip().lower() if row[12] else None,
                "bank": str(row[13]).strip().upper() if row[13] else None,
                "account": str(row[14]).strip() if row[14] else None,
            }
            final_data.append(worker)

    with open(OUTPUT_JSON, 'w', encoding='utf-8') as f:
        json.dump(final_data, f, indent=2, ensure_ascii=False)
    
    print(f"Standardization Complete. Processed {len(final_data)} workers.")
    print(f"File saved to {OUTPUT_JSON}")

except Exception as e:
    import traceback
    traceback.print_exc()
